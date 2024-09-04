from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.options import Options

from selenium.webdriver.common.by import By

from time import sleep
import random

import openpyxl

import PySimpleGUI as sg

workbook = openpyxl.load_workbook('produtos.xlsx')
sheet_produtos = workbook['Produtos']

dados = []

""" for linha in sheet_produtos.iter_rows(min_row=2):
    dados.append(linha[0].value)
    dados.append(linha[1].value)
    dados.append(linha[2].value)
    dados.append(linha[3].value)
    dados.append(linha[4].value)
    dados.append(linha[5].value)
    dados.append(linha[6].value)
    dados.append(linha[7].value)
    dados.append(linha[8].value)
    dados.append(linha[9].value)
    dados.append(linha[10].value)
    dados.append(linha[11].value)
    dados.append(linha[12].value)
    dados.append(linha[13].value)
    dados.append(linha[14].value)
    dados.append(linha[15].value)
    dados.append(linha[16].value) """

for linha in sheet_produtos.iter_rows(min_row=2, max_row=4):
    dados.append({'casa1' : linha[0].value,
             'casa2' : linha[1].value,
             'casa3' : linha[2].value,
             'casa4' : linha[3].value,
             'casa5' : linha[4].value,
             'casa6' : linha[5].value,
             'casa7' : linha[6].value,
             'casa8' : linha[7].value,
             'casa9' : linha[8].value,
             'casa10' : linha[9].value,
             'casa11' : linha[10].value,
             'casa12' : linha[11].value,
             'casa13' : linha[12].value,
             'casa14' : linha[13].value,
             'casa15' : linha[14].value,
             'casa16' : linha[15].value,
             'casa17' : linha[16].value,})

for dado in dados:
    print(dado['casa1'])

sg.theme('Reddit')
#sg.Push() serve para centralizar o elemento, para isto coloque em ambos os lados
layout = [
    [sg.Push(),sg.Button(button_text='Iniciar Automação'),sg.Push()]    
]

window = sg.Window('Inserir dados NF em site', layout)

edge_options = Options()
arguments = ['--lang=pt-BR', 'window-size=1000,1000', '--incognito']
for argument in arguments:
    edge_options.add_argument(argument)

caminho_padrao_para_download = 'C:\\Users\\rscop\\Desktop'

edge_options.add_experimental_option("prefs", {
    'download.default_directory': caminho_padrao_para_download,
    'download.directory_upgrade': True,
    'download.prompt_for_download': False,
    'profile.default_content_setting_values.notifications': 2,
    'profile.default_content_setting_values.automatic_downloads': 1
})

def digitar_naturalmente(nome_produto, campo_pesquisa):
    for letra in nome_produto:
        campo_pesquisa.send_keys(letra)
        sleep(random.randint(1,5)/30)

while True:
    event,values = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event =='Iniciar Automação':

        driver = webdriver.Edge(options=edge_options)
        driver.get('https://cadastro-produtos-devaprender.netlify.app/index.html')
        sleep(5)

        for dado in dados:

            nome_produto = driver.find_element(By.ID, "product_name")
            digitar_naturalmente(dado['casa1'], nome_produto)

            descricao = driver.find_element(By.ID, "description")
            digitar_naturalmente(dado['casa2'], descricao)

            categoria = driver.find_element(By.ID, "category")
            digitar_naturalmente(dado['casa3'], categoria)

            codigo_produto = driver.find_element(By.ID, "product_code")
            digitar_naturalmente(dado['casa4'], codigo_produto)

            peso = driver.find_element(By.ID, "weight")
            digitar_naturalmente(str(dado['casa5']), peso)

            dimensao = driver.find_element(By.ID, "dimensions")
            digitar_naturalmente(str(dado['casa6']), dimensao)

            btn_proximo = driver.find_element(By.XPATH, "//button[@class='btn btn-primary me-2']")
            driver.execute_script('arguments[0].click()', btn_proximo)

            sleep(4)

            preco = driver.find_element(By.ID, "price")
            digitar_naturalmente(str(dado['casa7']), preco)

            qtd_estoque = driver.find_element(By.ID, "stock")
            digitar_naturalmente(str(dado['casa8']), qtd_estoque)

            data_validade = driver.find_element(By.ID, "expiry_date")
            digitar_naturalmente(str(dado['casa9']), data_validade)

            cor = driver.find_element(By.ID, "color")
            digitar_naturalmente(str(dado['casa10']), cor)

            tamanho = driver.find_element(By.ID, "size")
            digitar_naturalmente(str(dado['casa11']), tamanho)

            material = driver.find_element(By.ID, "material")
            digitar_naturalmente(str(dado['casa12']), material)

            btn_proximo = driver.find_element(By.XPATH, "//button[@class='btn btn-primary me-2']")
            driver.execute_script('arguments[0].click()', btn_proximo)

            sleep(4)

            fabricante = driver.find_element(By.ID, "manufacturer")
            digitar_naturalmente(str(dado['casa13']), fabricante)

            pais = driver.find_element(By.ID, "country")
            digitar_naturalmente(str(dado['casa14']), pais)

            observacoes = driver.find_element(By.ID, "remarks")
            digitar_naturalmente(str(dado['casa15']), observacoes)

            codigo_barras = driver.find_element(By.ID, "barcode")
            digitar_naturalmente(str(dado['casa16']), codigo_barras)

            localizacao_armazen = driver.find_element(By.ID, "warehouse_location")
            digitar_naturalmente(str(dado['casa17']), localizacao_armazen)

            btn_proximo = driver.find_element(By.XPATH, "//button[@class='btn btn-primary me-2']")
            driver.execute_script('arguments[0].click()', btn_proximo)

            sleep(2)

            alerta = driver.switch_to.alert
            alerta.accept()

            sleep(4)
            
            btn_finalizacao = driver.find_element(By.XPATH, "//button[@class='btn btn-primary']")
            driver.execute_script('arguments[0].click()', btn_finalizacao)